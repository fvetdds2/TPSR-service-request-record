import streamlit as st
import pandas as pd
import plotly.express as px
import openpyxl
from datetime import datetime

st.set_page_config(page_title="TPSR Service Request Dashboard", layout="wide")

# ─────────────────────────────────────────
# Passcode Gate — must pass before anything else loads
# ─────────────────────────────────────────
PASSCODE = "TPSR2025"   # ← change this to your desired passcode

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.markdown(
        """
        <style>
        .lock-container {
            display: flex; flex-direction: column; align-items: center;
            justify-content: center; margin-top: 8vh;
        }
        .lock-title {
            font-size: 2rem; font-weight: 700; color: #007BFF; /* Primary brand blue */
            margin-bottom: 0.25rem;
        }
        .lock-subtitle {
            font-size: 1rem; color: #6C757D; /* Darker grey for subtitles */
            margin-bottom: 2rem;
        }
        </style>
        <div class="lock-container">
            <div style="font-size:3.5rem">🔒</div>
            <div class="lock-title">MMC TPSR Dashboard</div>
            <div class="lock-subtitle">Enter the passcode to continue</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    col_l, col_c, col_r = st.columns([1, 1, 1])
    with col_c:
        entered = st.text_input(
            "Passcode", type="password", placeholder="Enter passcode…",
            label_visibility="collapsed",
        )
        login_btn = st.button("🔓 Unlock", use_container_width=True)

        if login_btn:
            if entered == PASSCODE:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("❌ Incorrect passcode. Please try again.")
    st.stop()   

# ─────────────────────────────────────────
# Load & Clean Data
# ─────────────────────────────────────────
@st.cache_data
def load_data():
    wb = openpyxl.load_workbook("cost_recovery_record_from_2025.xlsx")
    ws = wb.active

    # Build header map from row 1: col_number -> header name
    header_map = {cell.column: cell.value for cell in ws[1] if cell.value}

    # Service columns = E(5) through L(12); M(13) = specie; N(14) = Cancer_Related_Project
    service_col_nums = [5, 6, 7, 8, 9, 10, 11, 12]
    service_cols     = [header_map[c] for c in service_col_nums]
    CANCER_COL       = 14  

    
    def excel_date_to_int(val):
        if isinstance(val, datetime):
            return (val - datetime(1899, 12, 30)).days
        return val

    records = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        if all(cell.value is None for cell in row):
            continue

        vals = {cell.column: cell.value for cell in row if cell.value is not None}

        name   = vals.get(1)
        date   = vals.get(2)
        status = vals.get(3)
        cost   = vals.get(4, 0) or 0

        svc = {}
        for col_num, col_name in zip(service_col_nums, service_cols):
            raw = vals.get(col_num)
            raw = excel_date_to_int(raw)
            svc[col_name] = float(raw) if isinstance(raw, (int, float)) else 0.0

        cancer_raw = vals.get(CANCER_COL, "")
        cancer_val = str(cancer_raw).strip().capitalize() if cancer_raw else "Unknown"
        if cancer_val not in ("Yes", "No"):
            cancer_val = "Unknown"

        records.append({
            "Requester_Name":         name,
            "Required_Date":          pd.to_datetime(date, errors="coerce"),
            "Status":                 status or "Unknown",
            "Cost_Recovery":          float(cost) if isinstance(cost, (int, float)) else 0.0,
            "Cancer_Related_Project": cancer_val,
            **svc,
        })

    df = pd.DataFrame(records)
    df["Month_Year"]       = df["Required_Date"].dt.to_period("M")
    df["Month_Year_Label"] = df["Required_Date"].dt.strftime("%b %Y")

    return df, service_cols


df, service_cols = load_data()

# Short display labels for charts
short_labels = {
    "FFPE processing & Embedding":               "FFPE Process & Embed",
    "FFPE sectioning & H&E stain":               "FFPE Section & H&E",
    "Frozen sectioning-unstained slide":         "Frozen Unstained",
    "Frozen sectioning & H&E stain":             "Frozen H&E",
    "Frozen sectioning-step section":            "Frozen Step Section",
    "Repository FFPE sectioning-unstained slide":"Repo FFPE Unstained",
    "histology tissue collection vials":         "Histology Vials",
    "histopathology support (hr)":               "Histopath Support (hr)",
}

# ─────────────────────────────────────────
# Color Palettes
# ─────────────────────────────────────────

# Status colors — used in BOTH sidebar labels and pie chart
STATUS_COLORS = {
    "Completed": "#28A745", # Green for success
    "Pending":   "#FFC107", # Orange/Yellow for warning/pending
    "Unknown":   "#6C757D", # Muted grey for unknown/neutral
}

# Requester colors — distinct color per requester, used in sidebar + bar charts
REQUESTER_PALETTE = [
    "#007BFF", "#6F42C1", "#DC3545", "#20C997",  # Primary brand blue, indigo, red, teal
    "#FD7E14", "#6610F2", "#17A2B8", "#E83E8C",  # Orange, purple, cyan, pink
    "#FFC107", "#28A745",                        # Yellow, green (can be used if distinct enough from status)
    "#343A40", "#A64D79"                         # Dark grey, a muted magenta for variety
]

all_requesters    = sorted(df["Requester_Name"].dropna().unique().tolist())
REQUESTER_COLORS  = {name: REQUESTER_PALETTE[i % len(REQUESTER_PALETTE)]
                     for i, name in enumerate(all_requesters)}

# ─────────────────────────────────────────
# Sidebar Filters
# ─────────────────────────────────────────
st.sidebar.header("🔍 Filters")

# --- Status filter with color-matched labels ---
st.sidebar.markdown("**Status**")
status_options  = sorted(df["Status"].unique().tolist())
status_filter   = []
for s in status_options:
    color = STATUS_COLORS.get(s, "#95a5a6") # Fallback to a neutral grey if not defined
    checked = st.sidebar.checkbox(
        label=s, value=True, key=f"status_{s}",
        help=f"Filter by {s}"
    )
    # Inject colored badge next to the checkbox via markdown
    st.sidebar.markdown(
        f'<span style="display:inline-block;background:{color};color:white;'
        f'padding:2px 10px;border-radius:12px;font-size:12px;'
        f'margin-bottom:4px">{s}</span>',
        unsafe_allow_html=True,
    )
    if checked:
        status_filter.append(s)

st.sidebar.markdown("---")

# --- Requester filter: all visible, each with its own color badge ---
st.sidebar.markdown("**Requester**")
requester_filter = []
for name in all_requesters:
    color = REQUESTER_COLORS[name]
    checked = st.sidebar.checkbox(label=name, value=True, key=f"req_{name}")
    st.sidebar.markdown(
        f'<span style="display:inline-block;background:{color};color:white;'
        f'padding:2px 10px;border-radius:12px;font-size:12px;'
        f'margin-bottom:4px;max-width:100%;word-break:break-word">{name}</span>',
        unsafe_allow_html=True,
    )
    if checked:
        requester_filter.append(name)

df_filtered = df[
    df["Status"].isin(status_filter) &
    df["Requester_Name"].isin(requester_filter)
]

# ─────────────────────────────────────────
# Title
# ─────────────────────────────────────────
st.title("MMC Translational Pathology Shared Resource Core Service Request Dashboard")
st.caption("Cost Recovery Record from April 2025 / Feb 2026")
st.divider()

# ─────────────────────────────────────────
# Metrics
# ─────────────────────────────────────────
completed      = int((df_filtered["Status"] == "Completed").sum())
pending        = int((df_filtered["Status"] == "Pending").sum())
total_cost     = df_filtered["Cost_Recovery"].sum()
total_services = int(df_filtered[service_cols].sum().sum())

c1, c2, c3, c4 = st.columns(4)
c1.metric("✅ Completed",           completed)
c2.metric("⏳ Pending",              pending)
c3.metric("💰 Total Cost Recovery",  f"${total_cost:,.2f}")
c4.metric("🧪 Total slides",  total_services)
st.divider()


# ─────────────────────────────────────────
# Cancer Related Project Chart
# ─────────────────────────────────────────
st.subheader("Cancer Related Project")

# Correctly unpack the single column from st.columns(1)
(cancer_col2,) = st.columns(1)

with cancer_col2:
    # Stacked bar: Cancer Yes/No per requester
    cancer_req = (
        df_filtered.groupby(["Requester_Name", "Cancer_Related_Project"])
        .size()
        .reset_index(name="Count")
    )

    fig_cancer_bar = px.bar(
        cancer_req,
        x="Requester_Name",
        y="Count",
        color="Cancer_Related_Project",
        barmode="stack",
        text="Count",
        color_discrete_map={
            "Yes": "#DC3545",
            "No": "#007BFF",
            "Unknown": "#6C757D",
        },  # Red for Yes, Blue for No, Grey for Unknown
        labels={
            "Requester_Name": "Requester",
            "Count": "Projects",
            "Cancer_Related_Project": "Cancer Related",
        },
    )

    fig_cancer_bar.update_traces(textposition="inside")
    fig_cancer_bar.update_layout(
        xaxis_tickangle=-30,
        legend_title_text="Cancer Related"
    )

    st.plotly_chart(fig_cancer_bar, use_container_width=True)

st.divider()


# ─────────────────────────────────────────
# Service Types Distribution (Cols E–L)
# ─────────────────────────────────────────
st.subheader("📊 Service Types Distribution")

svc_totals = df_filtered[service_cols].sum().reset_index()
svc_totals.columns = ["Service_Type", "Total_Units"]
svc_totals["Short_Label"] = svc_totals["Service_Type"].map(short_labels)
svc_totals = svc_totals.sort_values("Total_Units", ascending=False)

fig_svc = px.bar(
    svc_totals, x="Short_Label", y="Total_Units",
    color="Short_Label",
    color_discrete_sequence=px.colors.qualitative.Safe, # Good for distinct categories
    text="Total_Units",
    labels={"Short_Label": "Service Type", "Total_Units": "Total Units"},
)
fig_svc.update_traces(textposition="outside")
fig_svc.update_layout(showlegend=False, xaxis_title="Service Type", yaxis_title="Total Units")
st.plotly_chart(fig_svc, use_container_width=True)

# Per-requester grouped bar
st.subheader("Service Units per Requester by Type")
melt = (
    df_filtered[["Requester_Name"] + service_cols]
    .melt(id_vars="Requester_Name", value_vars=service_cols,
          var_name="Service_Type", value_name="Units")
)
melt["Service_Type"] = melt["Service_Type"].map(short_labels)
melt = melt[melt["Units"] > 0]

if not melt.empty:
    fig_grp = px.bar(
        melt, x="Requester_Name", y="Units",
        color="Service_Type", barmode="group", text="Units",
        color_discrete_sequence=px.colors.qualitative.Safe, # Good for distinct categories
        labels={"Requester_Name": "Requester", "Units": "Units", "Service_Type": "Service"},
    )
    fig_grp.update_traces(textposition="outside")
    fig_grp.update_layout(xaxis_tickangle=-30)
    st.plotly_chart(fig_grp, use_container_width=True)
else:
    st.info("No service unit data to display for the selected filters.")

# ─────────────────────────────────────────
# Total Service Count by Month
# ─────────────────────────────────────────
st.subheader("📅 Total Slide Count by Month")

svc_by_month = (
    df_filtered
    .assign(Total_Units=df_filtered[service_cols].sum(axis=1))
    .groupby(["Month_Year", "Month_Year_Label"], as_index=False)["Total_Units"]
    .sum()   # ← THIS is where .sum() goes
    .sort_values("Month_Year")
)

if not svc_by_month.empty and svc_by_month["Total_Units"].sum() > 0:
    fig_month = px.bar(
        svc_by_month,
        x="Month_Year_Label",
        y="Total_Units",
        color="Total_Units",
        color_continuous_scale="Blues",
        text="Total_Units",
        labels={
            "Month_Year_Label": "Month",
            "Total_Units": "Total Service Units"
        },
    )  # ← THIS closes px.bar()

    fig_month.update_traces(textposition="outside")
    fig_month.update_layout(
        xaxis_title="Month / Year",
        yaxis_title="Total Slides",
        coloraxis_showscale=False,
    )

    st.plotly_chart(fig_month, use_container_width=True)
else:
    st.info("No service unit data available for the selected filters.")
# ─────────────────────────────────────────
# Cost Recovery Over Time
# ─────────────────────────────────────────
st.subheader("💵 Cost Recovery Over Time")

cost_time = (
    df_filtered
    .groupby(["Month_Year", "Month_Year_Label"])["Cost_Recovery"]
    .sum()
    .reset_index()
    .sort_values("Month_Year")
)

if not cost_time.empty:
    fig_time = px.line(
        cost_time, x="Month_Year_Label", y="Cost_Recovery",
        markers=True, color_discrete_sequence=["#007BFF"], # Primary brand blue for line chart
        labels={"Month_Year_Label": "Month", "Cost_Recovery": "Cost Recovery ($)"},
    )
    fig_time.update_traces(line_width=2.5, marker_size=8)
    fig_time.update_layout(xaxis_title="Month / Year", yaxis_title="Cost ($)")
    st.plotly_chart(fig_time, use_container_width=True)
else:
    st.info("No date data available.")

# ─────────────────────────────────────────
# Raw Data Table
# ─────────────────────────────────────────
with st.expander("📋 Raw Data Table"):
    display_df = df_filtered[
        ["Requester_Name", "Month_Year_Label", "Status", "Cost_Recovery"] + service_cols
    ].copy()
    display_df = display_df.rename(columns={"Month_Year_Label": "Month / Year", **short_labels})
    st.dataframe(
        display_df.style.format({"Cost_Recovery": "${:,.2f}"}),
        use_container_width=True,
    )
